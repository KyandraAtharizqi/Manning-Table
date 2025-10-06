import pandas as pd
import io
from datetime import datetime
import streamlit as st

# Page config
st.set_page_config(page_title="Manning Table Generator", layout="wide")

# Title
st.title("📊 Manning Table Generator")
st.markdown("---")

# Session state
if 'cleaned_data' not in st.session_state:
    st.session_state.cleaned_data = None
if 'manning_table' not in st.session_state:
    st.session_state.manning_table = None

# Sidebar
with st.sidebar:
    st.header("📁 Upload Files")
    master_file = st.file_uploader("Upload MasterData Excel", type=['xlsx', 'xls'])
    structural_file = st.file_uploader("Upload StructuralMapping Excel", type=['xlsx', 'xls'])
    st.markdown("---")
    st.info("Upload both files to start processing")

# Tabs
tab1, tab2, tab3 = st.tabs(["📋 Cleaned Data", "📊 Manning Table", "ℹ️ Instructions"])

with tab3:
    st.header("How to Use")
    st.markdown("""
    ### Steps:
    1. Upload MasterData and StructuralMapping files
    2. Generate Cleaned Data (filters valid employees)
    3. Generate Manning Table (creates hierarchical structure with totals and breakdowns)
    
    ### Features:
    - Hierarchical grouping (Directorate Group → Directorate → Division → Department → Cost Center)
    - Automatic totals calculation at each level
    - Pangkat and Status breakdowns for Department, Directorate, and Directorate Group levels
    - Formatted Excel export matching VBA output
    """)

with tab1:
    st.header("Cleaned Data Generation")
    
    if master_file and structural_file:
        col1, col2 = st.columns([1, 4])
        
        with col1:
            if st.button("🔄 Generate Cleaned Data", type="primary"):
                with st.spinner("Processing..."):
                    try:
                        master_df = pd.read_excel(master_file)
                        structural_df = pd.read_excel(structural_file)
                        
                        # Filter valid employees
                        master_df['Reg. No.'] = master_df['Reg. No.'].astype(str).str.strip()
                        master_df['Nama'] = master_df['Nama'].astype(str).str.strip()
                        
                        valid_employees = master_df[
                            (master_df['Reg. No.'] != '') & 
                            (master_df['Reg. No.'] != '0') &
                            (master_df['Reg. No.'] != 'nan') &
                            (master_df['Nama'] != '') &
                            (master_df['Nama'] != 'nan')
                        ].copy().reset_index(drop=True)
                        
                        st.info(f"Found {len(valid_employees)} valid employees")
                        
                        # Build cleaned data - keep Position Code as text throughout
                        cleaned_records = []
                        
                        # Convert Position Code to text format (preserve leading zeros, etc.)
                        structural_df['Position Code'] = structural_df['Position Code'].apply(
                            lambda x: str(x).strip() if pd.notna(x) and str(x).strip() not in ['', 'nan'] else ''
                        )
                        
                        for _, emp in valid_employees.iterrows():
                            raw_pos_code = emp['Position Code']
                            # Keep as text, don't convert to number
                            if pd.isna(raw_pos_code) or str(raw_pos_code).strip() in ['', 'nan']:
                                pos_code = ''
                            else:
                                # Convert to string but preserve format (e.g., "0001" stays "0001")
                                pos_code = str(raw_pos_code).strip()
                                # If it looks like a number with leading zeros, keep them
                                if raw_pos_code != pos_code:
                                    pos_code = str(int(float(raw_pos_code))).zfill(len(str(raw_pos_code).split('.')[0]))
                            
                            if pos_code:
                                struct_match = structural_df[structural_df['Position Code'] == pos_code]
                                if len(struct_match) > 0:
                                    s = struct_match.iloc[0]
                                    pos_description = s['Position Name']
                                    pangkat_struktural = s['Level/Pangkat']
                                    directorate = s['Directorate']
                                    division = s['Division']
                                    department = s['Department']
                                    cost_center = s['CostCenter']
                                    dir_group = s['DepartmentGroup']
                                else:
                                    pos_description = pangkat_struktural = directorate = division = department = cost_center = dir_group = ''
                            else:
                                pos_description = pangkat_struktural = directorate = division = department = cost_center = dir_group = ''
                            
                            cleaned_records.append({
                                'Nama': emp['Nama'],
                                'Reg. No.': emp['Reg. No.'],
                                'Position Code': pos_code,
                                'Organization Description': emp['Organization Descrip'],
                                'Position Description': pos_description,
                                'Pangkat/Level': emp['Pangkat/Level'],
                                'Pangkat/Level Struktural': pangkat_struktural,
                                'Position Grade': emp['Grade'],
                                'Tgl. Mulai Bekerja': emp['Tgl. Mulai Bekerja'],
                                'Status': emp['Status'],
                                'Pendidikan': emp['Pendidikan'],
                                'Tgl. Pensiun': emp['Tgl. Pensiun'],
                                'Directorate': directorate,
                                'Divisione': division,
                                'Department': department,
                                'Cost Center': cost_center,
                                'Directorate Group': dir_group
                            })
                        
                        st.session_state.cleaned_data = pd.DataFrame(cleaned_records)
                        st.success(f"✅ Generated {len(cleaned_records)} employee records")
                        
                    except Exception as e:
                        st.error(f"Error: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())
        
        if st.session_state.cleaned_data is not None:
            st.subheader(f"Preview ({len(st.session_state.cleaned_data)} records)")
            st.dataframe(st.session_state.cleaned_data, use_container_width=True, height=400)
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                st.session_state.cleaned_data.to_excel(writer, index=False, sheet_name='CleanedData')
                
                # Format Position Code column as text to prevent conversion to numbers
                workbook = writer.book
                worksheet = writer.sheets['CleanedData']
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
                    for cell in row:
                        cell.number_format = '@'  # Text format
            
            output.seek(0)
            
            st.download_button(
                label="📥 Download Cleaned Data",
                data=output,
                file_name=f"CleanedData_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning("⚠️ Please upload both files")

with tab2:
    st.header("Manning Table Generation")
    
    use_uploaded = st.checkbox("Upload Cleaned Data Excel (instead of using generated)")
    cleaned_file = st.file_uploader("Upload CleanedData Excel", type=['xlsx', 'xls'], key="cleaned_upload") if use_uploaded else None
    
    has_cleaned = (st.session_state.cleaned_data is not None) or (cleaned_file is not None)
    
    if has_cleaned and structural_file:
        col1, col2 = st.columns([1, 4])
        
        with col1:
            if st.button("🔄 Generate Manning Table", type="primary"):
                with st.spinner("Building Manning Table..."):
                    try:
                        structural_df = pd.read_excel(structural_file)
                        cleaned_df = pd.read_excel(cleaned_file) if cleaned_file else st.session_state.cleaned_data.copy()
                        
                        # Clean position code function
                        def clean_position_code(x):
                            if pd.isna(x) or str(x).strip() in ['', 'nan']:
                                return ''
                            s = str(x).strip()
                            if '.' in s:
                                try:
                                    return str(int(float(s)))
                                except:
                                    return s.split('.')[0]
                            return s
                        
                        # Prepare data - clean position codes properly
                        structural_df['Position Code'] = structural_df['Position Code'].apply(clean_position_code)
                        structural_df['Level/Pangkat'] = structural_df['Level/Pangkat'].astype(str).str.strip()
                        
                        cleaned_df['Position Code'] = cleaned_df['Position Code'].apply(clean_position_code)
                        cleaned_df['Pangkat/Level'] = cleaned_df['Pangkat/Level'].astype(str).str.strip()
                        
                        # Track which employees have been assigned to prevent duplicates
                        assigned_employees = set()
                        
                        # Sort structural data
                        structural_df = structural_df.sort_values(
                            by=['DepartmentGroup', 'Directorate', 'Division', 'Department', 'CostCenter', 'Position Code', 'Level/Pangkat'],
                            na_position='last'
                        )
                        
                        # Build manning table with headers and totals
                        manning_records = []
                        
                        # Hierarchy trackers
                        last_dir_group = last_directorate = last_division = last_department = last_cc = ""
                        
                        # Totals accumulators
                        cc_std = cc_act = cc_vac = cc_exc = 0
                        dept_std = dept_act = dept_vac = dept_exc = 0
                        div_std = div_act = div_vac = div_exc = 0
                        dir_std = dir_act = dir_vac = dir_exc = 0
                        dirgrp_std = dirgrp_act = dirgrp_vac = dirgrp_exc = 0
                        
                        # Breakdown dictionaries
                        dept_level_dict = {}
                        dept_status_dict = {}
                        dir_level_dict = {}
                        dir_status_dict = {}
                        dirgrp_level_dict = {}
                        dirgrp_status_dict = {}
                        
                        def add_totals_row(level_name, name, std, act, vac, exc):
                            colors = {
                                'Cost Center': 'rgb(250, 235, 235)',
                                'Department': 'rgb(255, 245, 215)',
                                'Division': 'rgb(220, 240, 210)',
                                'Directorate': 'rgb(200, 220, 235)',
                                'Directorate Group': 'rgb(180, 200, 220)'
                            }
                            manning_records.append({
                                'ROW_TYPE': 'TOTAL',
                                'TOTAL_TEXT': f'TOTAL for {level_name}: {name}',
                                'TOTAL_COLOR': colors.get(level_name, 'rgb(235, 235, 235)'),
                                'Position Code': '', 'Position Description': '', 'Grade': '', 'Pangkat/Level': '',
                                'Standard': std if std != 0 else '',
                                'Actual': act,
                                'Vacant': vac,
                                'Excess': exc,
                                'Name': '', 'Reg. No.': '', 'Status': '', 'Education': '', 'Start Date': '', 'Retirement Date': ''
                            })
                        
                        def add_breakdown_rows(level_name, name, level_dict, status_dict):
                            if not level_dict and not status_dict:
                                return
                            
                            pangkat_text = "Pangkat breakdown: " + ", ".join([f"{k} ({v})" for k, v in level_dict.items()]) if level_dict else "Pangkat breakdown: -"
                            manning_records.append({
                                'ROW_TYPE': 'BREAKDOWN',
                                'BREAKDOWN_TEXT': pangkat_text,
                                'Position Code': '', 'Position Description': '', 'Grade': '', 'Pangkat/Level': '',
                                'Standard': '', 'Actual': '', 'Vacant': '', 'Excess': '',
                                'Name': '', 'Reg. No.': '', 'Status': '', 'Education': '', 'Start Date': '', 'Retirement Date': ''
                            })
                            
                            status_text = "Status breakdown: " + ", ".join([f"{k} ({v})" for k, v in status_dict.items()]) if status_dict else "Status breakdown: -"
                            manning_records.append({
                                'ROW_TYPE': 'BREAKDOWN',
                                'BREAKDOWN_TEXT': status_text,
                                'Position Code': '', 'Position Description': '', 'Grade': '', 'Pangkat/Level': '',
                                'Standard': '', 'Actual': '', 'Vacant': '', 'Excess': '',
                                'Name': '', 'Reg. No.': '', 'Status': '', 'Education': '', 'Start Date': '', 'Retirement Date': ''
                            })
                        
                        def add_header_row(text, color):
                            manning_records.append({
                                'ROW_TYPE': 'HEADER',
                                'HEADER_TEXT': text,
                                'HEADER_COLOR': color,
                                'Position Code': '', 'Position Description': '', 'Grade': '', 'Pangkat/Level': '',
                                'Standard': '', 'Actual': '', 'Vacant': '', 'Excess': '',
                                'Name': '', 'Reg. No.': '', 'Status': '', 'Education': '', 'Start Date': '', 'Retirement Date': ''
                            })
                        
                        for _, struct_row in structural_df.iterrows():
                            pos_code = str(struct_row['Position Code']).strip()
                            level = str(struct_row['Level/Pangkat']).strip()
                            standard = struct_row['Standard']
                            dir_group = str(struct_row.get('DepartmentGroup', '')).strip()
                            directorate = str(struct_row.get('Directorate', '')).strip()
                            division = str(struct_row.get('Division', '')).strip()
                            department = str(struct_row.get('Department', '')).strip()
                            cc = str(struct_row.get('CostCenter', '')).strip()
                            
                            # Print totals for completed groups
                            if last_cc != "" and cc != last_cc:
                                add_totals_row("Cost Center", last_cc, cc_std, cc_act, cc_vac, cc_exc)
                                cc_std = cc_act = cc_vac = cc_exc = 0
                            
                            if last_department != "" and department != last_department:
                                add_totals_row("Department", last_department, dept_std, dept_act, dept_vac, dept_exc)
                                add_breakdown_rows("Department", last_department, dept_level_dict, dept_status_dict)
                                dept_level_dict = {}
                                dept_status_dict = {}
                                dept_std = dept_act = dept_vac = dept_exc = 0
                            
                            if last_division != "" and division != last_division:
                                add_totals_row("Division", last_division, div_std, div_act, div_vac, div_exc)
                                div_std = div_act = div_vac = div_exc = 0
                            
                            if last_directorate != "" and directorate != last_directorate:
                                add_totals_row("Directorate", last_directorate, dir_std, dir_act, dir_vac, dir_exc)
                                add_breakdown_rows("Directorate", last_directorate, dir_level_dict, dir_status_dict)
                                dir_level_dict = {}
                                dir_status_dict = {}
                                dir_std = dir_act = dir_vac = dir_exc = 0
                            
                            if last_dir_group != "" and dir_group != last_dir_group:
                                add_totals_row("Directorate Group", last_dir_group, dirgrp_std, dirgrp_act, dirgrp_vac, dirgrp_exc)
                                add_breakdown_rows("Directorate Group", last_dir_group, dirgrp_level_dict, dirgrp_status_dict)
                                dirgrp_level_dict = {}
                                dirgrp_status_dict = {}
                                dirgrp_std = dirgrp_act = dirgrp_vac = dirgrp_exc = 0
                            
                            # Print headers for new groups
                            if dir_group != last_dir_group:
                                add_header_row(f"Directorate Group: {dir_group}", "rgb(180, 200, 220)")
                                last_dir_group = dir_group
                                last_directorate = last_division = last_department = last_cc = ""
                            
                            if directorate != last_directorate:
                                add_header_row(f"Directorate: {directorate}", "rgb(200, 220, 235)")
                                last_directorate = directorate
                                last_division = last_department = last_cc = ""
                            
                            if division != last_division:
                                add_header_row(f"Division: {division}", "rgb(220, 240, 210)")
                                last_division = division
                                last_department = last_cc = ""
                            
                            if department != last_department:
                                add_header_row(f"Department: {department}", "rgb(255, 245, 215)")
                                last_department = department
                                last_cc = ""
                            
                            if cc != last_cc:
                                add_header_row(f"Cost Center: {cc}", "rgb(250, 235, 235)")
                                last_cc = cc
                            
                            # Match employees - exclude already assigned ones
                            matching_emps = cleaned_df[
                                (cleaned_df['Position Code'] == pos_code) &
                                (cleaned_df['Pangkat/Level'] == level) &
                                (~cleaned_df['Reg. No.'].isin(assigned_employees))
                            ]
                            
                            actual = len(matching_emps)
                            
                            # Mark these employees as assigned
                            for _, emp in matching_emps.iterrows():
                                assigned_employees.add(emp['Reg. No.'])
                            
                            # Calculate metrics
                            if pd.isna(standard) or str(standard).strip() in ['*', '']:
                                vacant = excess = '*'
                                std_display = '*'
                                std_numeric = 0
                            else:
                                try:
                                    std_numeric = float(standard)
                                    std_display = std_numeric
                                    vacant = max(0, std_numeric - actual)
                                    excess = max(0, actual - std_numeric)
                                except:
                                    vacant = excess = '*'
                                    std_display = '*'
                                    std_numeric = 0
                            
                            # Update totals
                            if std_numeric > 0:
                                cc_std += std_numeric
                                dept_std += std_numeric
                                div_std += std_numeric
                                dir_std += std_numeric
                                dirgrp_std += std_numeric
                            
                            cc_act += actual
                            dept_act += actual
                            div_act += actual
                            dir_act += actual
                            dirgrp_act += actual
                            
                            if vacant != '*':
                                cc_vac += vacant
                                dept_vac += vacant
                                div_vac += vacant
                                dir_vac += vacant
                                dirgrp_vac += vacant
                            
                            if excess != '*':
                                cc_exc += excess
                                dept_exc += excess
                                div_exc += excess
                                dir_exc += excess
                                dirgrp_exc += excess
                            
                            # Update breakdowns
                            for _, emp in matching_emps.iterrows():
                                emp_level = str(emp.get('Pangkat/Level', '')).strip()
                                emp_status = str(emp.get('Status', '')).strip()
                                
                                if department:
                                    dept_level_dict[emp_level] = dept_level_dict.get(emp_level, 0) + 1
                                    dept_status_dict[emp_status] = dept_status_dict.get(emp_status, 0) + 1
                                
                                if directorate:
                                    dir_level_dict[emp_level] = dir_level_dict.get(emp_level, 0) + 1
                                    dir_status_dict[emp_status] = dir_status_dict.get(emp_status, 0) + 1
                                
                                if dir_group:
                                    dirgrp_level_dict[emp_level] = dirgrp_level_dict.get(emp_level, 0) + 1
                                    dirgrp_status_dict[emp_status] = dirgrp_status_dict.get(emp_status, 0) + 1
                            
                            # Build position record
                            base_record = {
                                'ROW_TYPE': 'DATA',
                                'Position Code': pos_code,
                                'Position Description': struct_row.get('Position Name', ''),
                                'Grade': struct_row.get('Grade', ''),
                                'Pangkat/Level': level,
                                'Standard': std_display,
                                'Actual': actual,
                                'Vacant': vacant,
                                'Excess': excess
                            }
                            
                            # Add employee rows
                            if actual > 0:
                                first = True
                                for _, emp in matching_emps.iterrows():
                                    if first:
                                        record = base_record.copy()
                                        record.update({
                                            'Name': emp.get('Nama', ''),
                                            'Reg. No.': emp.get('Reg. No.', ''),
                                            'Status': emp.get('Status', ''),
                                            'Education': emp.get('Pendidikan', ''),
                                            'Start Date': emp.get('Tgl. Mulai Bekerja', ''),
                                            'Retirement Date': emp.get('Tgl. Pensiun', '')
                                        })
                                        manning_records.append(record)
                                        first = False
                                    else:
                                        manning_records.append({
                                            'ROW_TYPE': 'DATA',
                                            'Position Code': '', 'Position Description': '', 'Grade': '', 'Pangkat/Level': '',
                                            'Standard': '', 'Actual': '', 'Vacant': '', 'Excess': '',
                                            'Name': emp.get('Nama', ''),
                                            'Reg. No.': emp.get('Reg. No.', ''),
                                            'Status': emp.get('Status', ''),
                                            'Education': emp.get('Pendidikan', ''),
                                            'Start Date': emp.get('Tgl. Mulai Bekerja', ''),
                                            'Retirement Date': emp.get('Tgl. Pensiun', '')
                                        })
                            else:
                                record = base_record.copy()
                                record.update({
                                    'Name': '–',
                                    'Reg. No.': '', 'Status': '', 'Education': '', 'Start Date': '', 'Retirement Date': ''
                                })
                                manning_records.append(record)
                        
                        # Final totals
                        if last_cc:
                            add_totals_row("Cost Center", last_cc, cc_std, cc_act, cc_vac, cc_exc)
                        if last_department:
                            add_totals_row("Department", last_department, dept_std, dept_act, dept_vac, dept_exc)
                            add_breakdown_rows("Department", last_department, dept_level_dict, dept_status_dict)
                        if last_division:
                            add_totals_row("Division", last_division, div_std, div_act, div_vac, div_exc)
                        if last_directorate:
                            add_totals_row("Directorate", last_directorate, dir_std, dir_act, dir_vac, dir_exc)
                            add_breakdown_rows("Directorate", last_directorate, dir_level_dict, dir_status_dict)
                        if last_dir_group:
                            add_totals_row("Directorate Group", last_dir_group, dirgrp_std, dirgrp_act, dirgrp_vac, dirgrp_exc)
                            add_breakdown_rows("Directorate Group", last_dir_group, dirgrp_level_dict, dirgrp_status_dict)
                        
                        manning_table = pd.DataFrame(manning_records)
                        st.session_state.manning_table = manning_table
                        st.success(f"✅ Manning Table generated! {len(manning_table)} records")
                        
                    except Exception as e:
                        st.error(f"Error: {str(e)}")
                        import traceback
                        st.code(traceback.format_exc())
        
        if st.session_state.manning_table is not None:
            st.subheader(f"Preview ({len(st.session_state.manning_table)} records)")
            
            
            # Display table
            display_df = st.session_state.manning_table.copy()
            display_df.loc[display_df['ROW_TYPE'] == 'HEADER', 'Position Code'] = display_df['HEADER_TEXT']
            display_df.loc[display_df['ROW_TYPE'] == 'TOTAL', 'Position Code'] = display_df['TOTAL_TEXT']
            display_df.loc[display_df['ROW_TYPE'] == 'BREAKDOWN', 'Position Code'] = display_df['BREAKDOWN_TEXT']
            
            display_cols = ['Position Code', 'Position Description', 'Grade', 'Pangkat/Level',
                          'Standard', 'Actual', 'Vacant', 'Excess',
                          'Name', 'Reg. No.', 'Status', 'Education', 'Start Date', 'Retirement Date']
            
            st.dataframe(display_df[display_cols], use_container_width=True, height=500)
            
            # Excel export with formatting
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            output = io.BytesIO()
            excel_df = st.session_state.manning_table.copy()
            
            for idx, row in excel_df.iterrows():
                if row['ROW_TYPE'] == 'HEADER':
                    excel_df.at[idx, 'Position Code'] = row['HEADER_TEXT']
                elif row['ROW_TYPE'] == 'TOTAL':
                    excel_df.at[idx, 'Position Code'] = row['TOTAL_TEXT']
                elif row['ROW_TYPE'] == 'BREAKDOWN':
                    excel_df.at[idx, 'Position Code'] = row['BREAKDOWN_TEXT']
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                excel_df[display_cols].to_excel(writer, sheet_name='ManningTable', startrow=3, index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['ManningTable']
                
                # Format all cells as text to prevent Excel from converting position codes to numbers
                for row in worksheet.iter_rows(min_row=5, max_row=worksheet.max_row, min_col=1, max_col=1):
                    for cell in row:
                        cell.number_format = '@'  # Text format
                
                # Title
                worksheet.merge_cells('A1:N1')
                title_cell = worksheet['A1']
                title_cell.value = 'MANNING TABLE'
                title_cell.font = Font(bold=True, size=18, color='FFFFFF')
                title_cell.fill = PatternFill(start_color='0F4C75', end_color='0F4C75', fill_type='solid')
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                worksheet.row_dimensions[1].height = 25
                
                # Header row
                header_fill = PatternFill(start_color='E1E6EB', end_color='E1E6EB', fill_type='solid')
                for col in range(1, 15):
                    cell = worksheet.cell(row=4, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                
                # Format data rows with section borders
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
                thick_top = Border(top=Side(style='thick'))
                
                # Define section separators (medium borders between logical groups)
                medium_right = Side(style='medium', color='000000')
                section_border_after_grade = Border(right=medium_right)  # After column D (Grade)
                section_border_after_excess = Border(right=medium_right)  # After column H (Excess)
                
                def hex_to_rgb(rgb_str):
                    if 'rgb' in rgb_str:
                        rgb = rgb_str.replace('rgb(', '').replace(')', '').split(',')
                        return '{:02X}{:02X}{:02X}'.format(int(rgb[0].strip()), int(rgb[1].strip()), int(rgb[2].strip()))
                    return 'FFFFFF'
                
                # Apply section borders to header row
                worksheet.cell(row=4, column=4).border = Border(
                    left=worksheet.cell(row=4, column=4).border.left,
                    right=medium_right,
                    top=worksheet.cell(row=4, column=4).border.top,
                    bottom=worksheet.cell(row=4, column=4).border.bottom
                )
                worksheet.cell(row=4, column=8).border = Border(
                    left=worksheet.cell(row=4, column=8).border.left,
                    right=medium_right,
                    top=worksheet.cell(row=4, column=8).border.top,
                    bottom=worksheet.cell(row=4, column=8).border.bottom
                )
                
                for idx, row in excel_df.iterrows():
                    excel_row = idx + 5
                    row_type = row['ROW_TYPE']
                    
                    if row_type == 'HEADER':
                        worksheet.merge_cells(f'A{excel_row}:N{excel_row}')
                        cell = worksheet[f'A{excel_row}']
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(start_color=hex_to_rgb(row['HEADER_COLOR']),
                                              end_color=hex_to_rgb(row['HEADER_COLOR']),
                                              fill_type='solid')
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    elif row_type == 'TOTAL':
                        color = hex_to_rgb(row['TOTAL_COLOR'])
                        for col in range(1, 15):
                            cell = worksheet.cell(row=excel_row, column=col)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                            
                            # Base border
                            left_border = Side(style='thin')
                            right_border = Side(style='thin')
                            top_border = Side(style='thin')
                            bottom_border = Side(style='thin')
                            
                            # Thick top border for numeric columns
                            if 5 <= col <= 8:
                                top_border = Side(style='thick')
                            
                            # Section separators
                            if col == 4:
                                right_border = medium_right
                            elif col == 8:
                                right_border = medium_right
                            
                            cell.border = Border(left=left_border, right=right_border, 
                                               top=top_border, bottom=bottom_border)
                    
                    elif row_type == 'BREAKDOWN':
                        worksheet.merge_cells(f'A{excel_row}:N{excel_row}')
                        cell = worksheet[f'A{excel_row}']
                        cell.font = Font(italic=True, size=10, color='505050')
                        cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    elif row_type == 'DATA':
                        # Apply section borders to data rows
                        # Section 1: Position Info (A-D) | Section 2: Numbers (E-H) | Section 3: Employee Info (I-N)
                        for col in range(1, 15):
                            cell = worksheet.cell(row=excel_row, column=col)
                            cell.border = thin_border
                            
                            # Add medium border after Grade (col D)
                            if col == 4:
                                cell.border = Border(
                                    left=cell.border.left,
                                    right=medium_right,
                                    top=cell.border.top,
                                    bottom=cell.border.bottom
                                )
                            # Add medium border after Excess (col H)
                            elif col == 8:
                                cell.border = Border(
                                    left=cell.border.left,
                                    right=medium_right,
                                    top=cell.border.top,
                                    bottom=cell.border.bottom
                                )
                
                # Add alternating row colors for better readability (only for DATA rows)
                for idx, row in excel_df.iterrows():
                    excel_row = idx + 5
                    if row['ROW_TYPE'] == 'DATA':
                        # Very light alternating colors
                        if (idx % 2) == 0:
                            light_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
                        else:
                            light_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                        
                        for col in range(1, 15):
                            cell = worksheet.cell(row=excel_row, column=col)
                            if not cell.value or cell.value == '':
                                cell.fill = light_fill
                
                # Center align numeric columns
                for row_idx in range(5, worksheet.max_row + 1):
                    for col in [3, 4, 5, 6, 7, 8]:  # Grade, Pangkat/Level, Standard, Actual, Vacant, Excess
                        cell = worksheet.cell(row=row_idx, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Left align text columns
                for row_idx in range(5, worksheet.max_row + 1):
                    for col in [1, 2, 9, 10, 11, 12, 13, 14]:  # Position info and employee info
                        cell = worksheet.cell(row=row_idx, column=col)
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                worksheet.column_dimensions['A'].width = 14
                worksheet.column_dimensions['B'].width = 30
                worksheet.column_dimensions['C'].width = 10
                worksheet.column_dimensions['D'].width = 10
                for col in ['E', 'F', 'G', 'H']:
                    worksheet.column_dimensions[col].width = 9
                for col in ['I', 'J', 'K', 'L', 'M', 'N']:
                    worksheet.column_dimensions[col].width = 16
            
            output.seek(0)
            
            st.download_button(
                label="📥 Download Manning Table (Excel)",
                data=output,
                file_name=f"ManningTable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    elif st.session_state.cleaned_data is None and not cleaned_file:
        st.warning("⚠️ Please generate Cleaned Data first or upload a Cleaned Data file")
    else:
        st.warning("⚠️ Please upload StructuralMapping file")

st.markdown("---")
st.markdown("Manning Table System v2.0 | Built with Streamlit")